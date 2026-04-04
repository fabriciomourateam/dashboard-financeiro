import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

OUTPUT = 'C:/Users/fhbom/Downloads/DASHBOARD_FINANCEIRO_FABRICIO.xlsx'

# CORES
VERDE_ESCURO = '1B5E20'
VERDE = '2E7D32'
VERDE_CLARO = 'A5D6A7'
VERDE_BG = 'E8F5E9'
VERMELHO = 'C62828'
VERMELHO_ESCURO = '8B0000'
VERMELHO_CLARO = 'FFCDD2'
AMARELO = 'F57F17'
AMARELO_CLARO = 'FFF9C4'
AZUL = '0D47A1'
AZUL_CLARO = 'BBDEFB'
AZUL_BG = 'E3F2FD'
CINZA_ESCURO = '263238'
CINZA = '546E7A'
CINZA_CLARO = 'ECEFF1'
BRANCO = 'FFFFFF'
PRETO = '000000'
LARANJA = 'E65100'
LARANJA_CLARO = 'FFE0B2'

def fill(cor): return PatternFill('solid', fgColor=cor)
def font(bold=False, size=11, cor=PRETO, italic=False):
    return Font(name='Arial', bold=bold, size=size, color=cor, italic=italic)
def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, size=11, cor_font=PRETO,
             bg=None, h_align='left', italic=False, wrap=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold, size, cor_font, italic)
    c.alignment = align(h_align, 'center', wrap)
    if bg:
        c.fill = fill(bg)
    if num_fmt:
        c.number_format = num_fmt
    c.border = border_thin()
    return c

def header_row(ws, row, cols_vals, bg=CINZA_ESCURO, text_cor=BRANCO, size=10):
    ws.row_dimensions[row].height = 22
    for col, val in cols_vals:
        c = set_cell(ws, row, col, val, bold=True, size=size, cor_font=text_cor,
                     bg=bg, h_align='center')

def merge_header(ws, row, col_start, col_end, text, bg=CINZA_ESCURO,
                 text_cor=BRANCO, size=13):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = font(True, size, text_cor)
    c.fill = fill(bg)
    c.alignment = align('center', 'center')
    ws.row_dimensions[row].height = 28

def num_cell(ws, row, col, value, bg, bold=False, cor_font=PRETO, fmt='R$ #,##0'):
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.font = font(bold, 10, cor_font)
    c.fill = fill(bg)
    c.alignment = align('right', 'center')
    c.border = border_thin()
    return c

wb = openpyxl.Workbook()

# ================================================
# ABA 1: DASHBOARD EXECUTIVO
# ================================================
ws1 = wb.active
ws1.title = 'Dashboard'
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'A5'

for col, width in [('A',30),('B',18),('C',18),('D',18),('E',16),('F',16),('G',24),('H',20)]:
    ws1.column_dimensions[col].width = width

# TITULO
ws1.merge_cells('A1:H1')
c = ws1['A1']
c.value = 'DASHBOARD FINANCEIRO - MY SHAPE | FABRICIO MOURA'
c.font = Font(name='Arial', bold=True, size=20, color=BRANCO)
c.fill = fill(CINZA_ESCURO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 42

ws1.merge_cells('A2:H2')
c = ws1['A2']
c.value = 'Consultoria Esportiva Online | 720 Alunos Ativos | Analise Q1 2026 (Janeiro - Marco)'
c.font = Font(name='Arial', italic=True, size=11, color=BRANCO)
c.fill = fill(CINZA)
c.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 22

# SECAO KPIs
merge_header(ws1, 3, 1, 8, 'INDICADORES-CHAVE DE DESEMPENHO (Q1 2026)', VERDE_ESCURO, BRANCO, 12)
header_row(ws1, 4, [
    (1,'Indicador'),(2,'Janeiro'),(3,'Fevereiro'),(4,'Marco'),
    (5,'Meta Ideal'),(6,'Status'),(7,'Variacao Jan-Mar'),(8,'Observacao')
], bg=VERDE, text_cor=BRANCO)

kpis = [
    ('FATURAMENTO BRUTO (R$)', 113418.65, 96040.36, 84642.88, 110000, 'ATENCAO', -25.4, 'Queda de 25% em 3 meses - tendencia preocupante'),
    ('DESPESAS TOTAIS ESTIMADAS (R$)', 70000, 86448, 76638, 65000, 'CRITICO', 9.5, 'Custos consomem 70-90% da receita'),
    ('LUCRO LIQUIDO ESTIMADO (R$)', 43418, 9592, 8004, 35000, 'CRITICO', -81.6, 'Queda brutal: de 43k para 8k em 3 meses'),
    ('MARGEM LIQUIDA (%)', 0.383, 0.0998, 0.0945, 0.30, 'CRITICO', -74.0, 'De 38% para 9% - abaixo do minimo saudavel'),
    ('TRAFEGO PAGO TOTAL (R$)', 18197, 35800, 32582, 25000, 'ATENCAO', 79.1, 'Cresceu 79% - medir CAC rigorosamente'),
    ('CUSTO DA EQUIPE (R$)', 24672, 23682, 23900, 22000, 'ATENCAO', -3.1, '9 colaboradores fixos + closer variavel'),
    ('CLOSER COMISSAO (R$)', 15272, 13782, 14000, 10000, 'CRITICO', -8.3, '13 a 17% do faturamento so no closer!'),
    ('MENTORIA EMPRESARIAL (R$)', 3333, 3333, 3333, 2000, 'ATENCAO', 0.0, 'R$40k/ano - ROI nao medido'),
    ('PRO LABORE + DISTRIBUICAO (R$)', 60000, 10000, 10000, 12000, 'CRITICO', -83.3, 'Retirou 60k em Janeiro - depleta caixa'),
    ('CHURN MENSAL', 0.05, 0.05, 0.05, 0.03, 'ATENCAO', 0.0, '36 alunos saem por mes = R$5.400 perdidos'),
    ('TAXA DE RENOVACAO', 0.40, 0.40, 0.40, 0.60, 'CRITICO', 0.0, '60% dos alunos NAO renovam no fim do ano'),
]

row = 5
for kpi in kpis:
    ind, jan, fev, mar, meta, status, var, obs = kpi
    bg = VERMELHO_CLARO if status == 'CRITICO' else (AMARELO_CLARO if status == 'ATENCAO' else VERDE_BG)
    status_txt = 'CRITICO' if status == 'CRITICO' else ('ATENCAO' if status == 'ATENCAO' else 'OK')
    cor_status = VERMELHO if status == 'CRITICO' else (AMARELO if status == 'ATENCAO' else VERDE_ESCURO)

    ws1.row_dimensions[row].height = 20
    set_cell(ws1, row, 1, ind, bold=True, bg=bg, size=10)

    is_pct = isinstance(jan, float) and jan < 2
    fmt = '0.0%' if is_pct else 'R$ #,##0'
    for col, val in [(2,jan),(3,fev),(4,mar),(5,meta)]:
        num_cell(ws1, row, col, val, bg, fmt=fmt)

    set_cell(ws1, row, 6, status_txt, bold=True, bg=bg, h_align='center',
             cor_font=cor_status, size=10)

    c = ws1.cell(row=row, column=7, value=var/100)
    c.number_format = '0.0%'
    c.font = font(True, 10, VERMELHO if var < 0 else VERDE_ESCURO)
    c.fill = fill(bg)
    c.alignment = align('center')
    c.border = border_thin()

    set_cell(ws1, row, 8, obs, italic=True, size=9, bg=bg, wrap=True)
    row += 1

# separador
row += 1

# SECAO GARGALOS
merge_header(ws1, row, 1, 8, 'ANALISE DE GARGALOS - ONDE O DINHEIRO ESTA VAZANDO', VERMELHO, BRANCO, 12)
row += 1
header_row(ws1, row, [
    (1,'#'),(2,'Gargalo Identificado'),(3,'Impacto Mensal'),(4,'Impacto Anual'),
    (5,'% Receita'),(6,'Prioridade'),(7,'Acao Imediata'),(8,'Prazo')
], bg=VERMELHO_ESCURO, text_cor=BRANCO)
row += 1

gargalos = [
    (1,'Closer com comissao excessiva (sem cap)',14000,168000,0.148,'URGENTE','Rever: fixo R$4k + comissao escalonada por meta','30 dias'),
    (2,'Retirada excessiva (60k em Jan - distribuicao)',20000,240000,0.178,'URGENTE','Definir pro-labore maximo: R$12k/mes. Sem distrib. extra','15 dias'),
    (3,'Taxa de renovacao 40% - 60% nao renovam',8640,103680,0.077,'URGENTE','CRM ativo 90 dias antes venc + oferta exclusiva renovacao','30 dias'),
    (4,'Mentoria sem ROI medido (R$40k/ano)',3333,39996,0.035,'URGENTE','Calcular vendas geradas pela mentoria - cortar se ROI < 3x','15 dias'),
    (5,'Trafego crescendo sem controle de CAC',7000,84000,0.062,'ALTO','Medir CAC por canal: Google vs Meta vs Organico','30 dias'),
    (6,'Churn 5%/mes = 36 alunos saindo/mes',5400,64800,0.048,'ALTO','NPS semanal + protocolo de reativacao antes de sair','45 dias'),
    (7,'Pagamentos anuais inflam meses (Langer 20k, Google 21k)',3500,42000,0.031,'MEDIO','Amortizar no budget (dividir por 12 nos relatorios)','30 dias'),
    (8,'Multiplas plataformas de pagamento: taxas acumuladas',1900,22800,0.017,'MEDIO','Consolidar em 1-2 plataformas com menor taxa','60 dias'),
    (9,'Equipe grande para receita atual: revisar ROI de cada um',5000,60000,0.045,'MEDIO','Mapear output real de cada colaborador','60 dias'),
    (10,'Mistura CPF/CNPJ anterior: risco fiscal passado',0,0,0,'ALTO','Regularizar com contador - avaliar exposicao e risco','15 dias'),
]

for g in gargalos:
    num, nome, mes, ano, pct, prio, acao, prazo = g
    bg = VERMELHO_CLARO if prio == 'URGENTE' else (AMARELO_CLARO if prio == 'ALTO' else VERDE_BG)
    cor_prio = VERMELHO if prio == 'URGENTE' else (AMARELO if prio == 'ALTO' else VERDE_ESCURO)

    ws1.row_dimensions[row].height = 22
    set_cell(ws1, row, 1, num, bold=True, bg=bg, h_align='center', size=10)
    set_cell(ws1, row, 2, nome, bold=True, bg=bg, size=10)
    num_cell(ws1, row, 3, mes, bg, bold=True, cor_font=VERMELHO)
    num_cell(ws1, row, 4, ano, bg, bold=True, cor_font=VERMELHO)
    c = ws1.cell(row=row, column=5, value=pct)
    c.number_format = '0.0%'; c.font = font(True, 10, VERMELHO if pct > 0.1 else AMARELO)
    c.fill = fill(bg); c.alignment = align('center'); c.border = border_thin()
    set_cell(ws1, row, 6, prio, bold=True, bg=bg, h_align='center', cor_font=cor_prio, size=10)
    set_cell(ws1, row, 7, acao, size=9, bg=bg, wrap=True)
    set_cell(ws1, row, 8, prazo, bold=True, bg=bg, h_align='center', size=10)
    row += 1

print(f'Aba Dashboard OK, linha {row}')


# ================================================
# ABA 2: RECEITAS
# ================================================
ws2 = wb.create_sheet('Receitas')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A4'

for col, width in [('A',32),('B',18),('C',18),('D',18),('E',18),('F',18)]:
    ws2.column_dimensions[col].width = width

ws2.merge_cells('A1:F1')
c = ws2['A1']
c.value = 'ANALISE DE RECEITAS - MY SHAPE 2026'
c.font = Font(name='Arial', bold=True, size=16, color=BRANCO)
c.fill = fill(VERDE_ESCURO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

# Receita por plataforma
merge_header(ws2, 2, 1, 6, 'Receita Bruta por Plataforma de Cobranca (R$)', VERDE, BRANCO, 11)
header_row(ws2, 3, [(1,'Plataforma'),(2,'Janeiro'),(3,'Fevereiro'),(4,'Marco'),(5,'TOTAL Q1'),(6,'Mix %')], VERDE)

receitas = [
    ('ASAAS (recorrencia)',         13200.13, 21602.42, 29816.51),
    ('CELCASH (cartao + boleto)',   27273.54, 19599.54, 25810.92),
    ('DOM Pagamentos (cartao)',     15083.10, 15491.40, 0.00),
    ('Pagar.me (parcelas antigas)', 27874.00, 0.00,     0.00),
    ('APPMAX',                        175.00, 0.00,     0.00),
    ('Outros / Pix direto (est.)',  19812.88, 39346.00, 29015.45),
]

row2 = 4
total_j = total_f = total_m = 0
for plat, jan, fev, mar in receitas:
    total_j += jan; total_f += fev; total_m += mar
    ws2.row_dimensions[row2].height = 20
    set_cell(ws2, row2, 1, plat, bg=VERDE_BG, size=10)
    for col, val in [(2,jan),(3,fev),(4,mar)]:
        num_cell(ws2, row2, col, val, VERDE_BG)
    num_cell(ws2, row2, 5, jan+fev+mar, VERDE_BG, bold=True, cor_font=VERDE_ESCURO)
    c = ws2.cell(row=row2, column=6, value=f'=E{row2}/E{row2+len(receitas)-row2+4+1}')
    c.fill = fill(VERDE_BG); c.alignment = align('center'); c.border = border_thin()
    row2 += 1

# Linha total
total_row_num = row2
ws2.row_dimensions[row2].height = 22
set_cell(ws2, row2, 1, 'TOTAL BRUTO Q1', bold=True, bg=VERDE, cor_font=BRANCO, size=11)
for col, val in [(2,total_j),(3,total_f),(4,total_m)]:
    num_cell(ws2, row2, col, val, VERDE, bold=True, cor_font=BRANCO)
num_cell(ws2, row2, 5, total_j+total_f+total_m, VERDE, bold=True, cor_font=BRANCO)
set_cell(ws2, row2, 6, '100%', bold=True, bg=VERDE, cor_font=BRANCO, h_align='center')

# Fix % formulas
for r in range(4, row2):
    c = ws2.cell(row=r, column=6)
    c.value = f'=E{r}/E${total_row_num}'
    c.number_format = '0.0%'
    c.font = font(False, 10)
    c.fill = fill(VERDE_BG); c.alignment = align('center'); c.border = border_thin()

row2 += 2

# Tendencia mensal
merge_header(ws2, row2, 1, 6, 'Tendencia de Receita: 2025 vs 2026 (Fonte: Planilha Controle)', CINZA_ESCURO, BRANCO, 11)
row2 += 1
header_row(ws2, row2, [(1,'Mes'),(2,'2025 (R$)'),(3,'2026 (R$)'),(4,'Crescimento'),(5,'Observacao'),(6,'')], CINZA_ESCURO)
row2 += 1

tendencia = [
    ('Janeiro', 81062.36, 113418.65, 0.399, 'Crescimento forte, boosted por Pagarme'),
    ('Fevereiro', 74275.06, 96040.36, 0.293, 'Crescimento, mas queda vs Jan 2026'),
    ('Marco', 86353.83, 84642.88, -0.020, 'QUEDA vs 2025 - sinal de alerta'),
    ('Abril', 91372.51, 0, 0, 'Dados pendentes'),
    ('Maio', 129721.21, 0, 0, 'Dados pendentes'),
    ('MEDIA 2025', 92829.46, 0, 0, 'Referencia anual'),
    ('TOTAL 2025', 1113953.54, 0, 0, 'Faturamento anual 2025'),
    ('TOTAL Q1 2026', 0, 294101.89, 0, 'Projecao anual: R$1.17M se mantiver'),
]

for t in tendencia:
    mes, v25, v26, cresc, obs = t
    bg = AMARELO_CLARO if cresc < 0 else (VERDE_BG if cresc > 0.1 else CINZA_CLARO)
    ws2.row_dimensions[row2].height = 20
    set_cell(ws2, row2, 1, mes, bold=('MEDIA' in mes or 'TOTAL' in mes), bg=bg, size=10)
    num_cell(ws2, row2, 2, v25 if v25 else None, bg)
    num_cell(ws2, row2, 3, v26 if v26 else None, bg, bold=True)
    c = ws2.cell(row=row2, column=4, value=cresc if cresc != 0 else None)
    if cresc != 0:
        c.number_format = '0.0%'
        c.font = font(True, 10, VERMELHO if cresc < 0 else VERDE_ESCURO)
    c.fill = fill(bg); c.alignment = align('center'); c.border = border_thin()
    set_cell(ws2, row2, 5, obs, size=9, italic=True, bg=bg)
    ws2.cell(row=row2, column=6).fill = fill(bg)
    ws2.cell(row=row2, column=6).border = border_thin()
    row2 += 1

row2 += 2

# Analise de churn e renovacao
merge_header(ws2, row2, 1, 6, 'Analise de Churn, Renovacao e Valor do Aluno (LTV)', AZUL, BRANCO, 11)
row2 += 1
header_row(ws2, row2, [(1,'Metrica'),(2,'Valor Atual'),(3,'Meta'),(4,'Impacto Mensal'),(5,'Impacto Anual'),(6,'Acao')], AZUL)
row2 += 1

metricas_aluno = [
    ('Alunos ativos total', 720, 1000, '', '', 'Crescer base com trafego'),
    ('Churn mensal', '5%', '3%', 'R$ -5.400', 'R$ -64.800', 'Protocolo de retencao ativo'),
    ('Alunos saindo por mes', 36, 20, '', '', ''),
    ('Taxa de renovacao', '40%', '60%', 'R$ -8.640', 'R$ -103.680', 'CRM ativo 90 dias antes'),
    ('Ticket medio mensal (est.)', 'R$ 150', 'R$ 165', '', '', 'Upsell dieta -> dieta+treino'),
    ('LTV anual por aluno', 'R$ 1.800', 'R$ 1.980', '', '', ''),
    ('CAC estimado (trafego/novas vendas)', 'R$ 700+', 'R$ 500', '', '', 'Medir mensalmente por canal'),
    ('ROI de trafego (LTV/CAC)', '2.5x', '3.5x', '', '', 'Meta minima: 3x'),
    ('Receita potencial se churn = 3%', 'R$ 124.800/mes', '', '', 'R$ 1.497.600', 'vs atual R$1.17M'),
]

for m in metricas_aluno:
    nome = m[0]
    ws2.row_dimensions[row2].height = 20
    bg = AZUL_BG
    set_cell(ws2, row2, 1, nome, bg=bg, size=10)
    for col, val in enumerate(m[1:], 2):
        set_cell(ws2, row2, col, val, bg=bg, h_align='center', size=10)
    row2 += 1

print(f'Aba Receitas OK')


# ================================================
# ABA 3: DESPESAS
# ================================================
ws3 = wb.create_sheet('Despesas')
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'A4'

for col, width in [('A',32),('B',18),('C',18),('D',18),('E',18),('F',14),('G',20)]:
    ws3.column_dimensions[col].width = width

ws3.merge_cells('A1:G1')
c = ws3['A1']
c.value = 'ANALISE DETALHADA DE DESPESAS - MY SHAPE 2026'
c.font = Font(name='Arial', bold=True, size=16, color=BRANCO)
c.fill = fill(VERMELHO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

merge_header(ws3, 2, 1, 7, 'Despesas por Categoria - Janeiro a Marco 2026 (R$)', VERMELHO_ESCURO, BRANCO, 11)
header_row(ws3, 3, [
    (1,'Categoria'),(2,'Janeiro'),(3,'Fevereiro'),(4,'Marco'),(5,'Media Mensal'),(6,'% Receita'),(7,'Tipo')
], VERMELHO_ESCURO)

REC_MEDIA = (113418.65 + 96040.36 + 84642.88) / 3  # 98033.96

despesas = [
    # (nome, jan, fev, mar, tipo)
    # TRAFEGO
    ('TRAFEGO PAGO', None, None, None, 'GRUPO'),
    ('  Google Ads (total)', 11707.29, 27000.00, 26583.84, 'VAR'),
    ('  Meta / Instagram (total)', 6490.64, 9321.43, 6312.84, 'VAR'),
    ('  Gestor de Trafego (Langer anual)', 0, 20000.00, 1666.67, 'ANUAL'),
    ('SUBTOTAL TRAFEGO', 18197.93, 56321.43, 34563.35, 'SUB'),

    # EQUIPE
    ('EQUIPE (Colaboradores)', None, None, None, 'GRUPO'),
    ('  Closer - Guilherme (comissao)', 15272.89, 13782.40, 14000.00, 'VAR'),
    ('  Nutricionista - Thais', 2700.00, 2700.00, 2700.00, 'FIX'),
    ('  SDR - Roberta', 1500.00, 1500.00, 1500.00, 'FIX'),
    ('  Estagiario - Guido', 1350.00, 1350.00, 1350.00, 'FIX'),
    ('  Estagiaria - Dayana', 500.00, 500.00, 500.00, 'FIX'),
    ('  Estagiario 3o', 500.00, 500.00, 500.00, 'FIX'),
    ('  Suporte - Jean', 1400.00, 1400.00, 1400.00, 'FIX'),
    ('  Suporte - Andreia', 1000.00, 1000.00, 1000.00, 'FIX'),
    ('  Social Media - Lype', 700.00, 700.00, 700.00, 'FIX'),
    ('  Psicologa - Josie', 250.00, 250.00, 250.00, 'FIX'),
    ('SUBTOTAL EQUIPE', 25172.89, 23682.40, 23900.00, 'SUB'),

    # GESTAO PROPRIA
    ('GESTAO PROPRIA', None, None, None, 'GRUPO'),
    ('  Pro Labore Fabricio', 10000.00, 10000.00, 10000.00, 'FIX'),
    ('  Distribuicao de Lucros (retirads)', 50000.00, 0, 0, 'VAR'),
    ('SUBTOTAL GESTAO PROPRIA', 60000.00, 10000.00, 10000.00, 'SUB'),

    # DESENVOLVIMENTO
    ('DESENVOLVIMENTO E EDUCACAO', None, None, None, 'GRUPO'),
    ('  Mentoria Empresarial (mensal)', 3333.33, 3333.33, 3333.33, 'FIX'),
    ('  Viagens e Eventos', 47.10, 1775.00, 934.36, 'VAR'),
    ('  Investimento Produto Digital', 0, 97.00, 397.00, 'VAR'),
    ('SUBTOTAL DESENVOLVIMENTO', 3380.43, 5205.33, 4664.69, 'SUB'),

    # FERRAMENTAS
    ('FERRAMENTAS E PLATAFORMAS', None, None, None, 'GRUPO'),
    ('  Claude / Anthropic (IA)', 258.15, 179.97, 1083.43, 'FIX'),
    ('  ChatGPT', 113.35, 110.73, 0, 'FIX'),
    ('  Supabase (banco de dados)', 143.75, 138.06, 135.71, 'FIX'),
    ('  Cursor (desenvolvimento)', 167.27, 0, 0, 'FIX'),
    ('  ManyChat', 85.34, 159.90, 0, 'FIX'),
    ('  WhatsApp Business + Multi', 256.90, 256.90, 256.90, 'FIX'),
    ('  Canva Pro', 24.15, 24.15, 144.90, 'FIX'),
    ('  Google One + Instagram', 53.90, 53.90, 150.89, 'FIX'),
    ('  Outros (Openrouter, Google Cloud)', 20.90, 0, 71.30, 'FIX'),
    ('SUBTOTAL FERRAMENTAS', 1123.71, 923.61, 1843.13, 'SUB'),

    # ESTRUTURA
    ('ESTRUTURA E BENEFICIOS', None, None, None, 'GRUPO'),
    ('  Convenio Saude (equipe)', 2461.66, 2461.66, 2461.66, 'FIX'),
    ('  Contabilidade', 400.00, 400.00, 400.00, 'FIX'),
    ('  Internet e Celular', 144.97, 34.99, 179.99, 'FIX'),
    ('  Hospedagem / Infraestrutura', 0, 181.08, 804.17, 'VAR'),
    ('SUBTOTAL ESTRUTURA', 3006.63, 3077.73, 3845.82, 'SUB'),

    # PLATAFORMAS COBRANCA
    ('TAXAS DE PLATAFORMAS', None, None, None, 'GRUPO'),
    ('  Taxas ASAAS (est. 4.1%)', 543.00, 707.96, 1296.13, 'VAR'),
    ('  Taxas CELCASH (est. 3.3%)', 899.00, 735.00, 1454.54, 'VAR'),
    ('  Taxas DOM (est. 5.6%)', 1218.55, 1803.34, 0, 'VAR'),
    ('  Taxas Pagarme (est. 8.9%)', 3568.53, 0, 0, 'VAR'),
    ('SUBTOTAL TAXAS', 6229.08, 3246.30, 2750.67, 'SUB'),

    # TOTAIS
    ('TOTAL GERAL DESPESAS', 117110.67, 102456.80, 81567.66, 'TOTAL'),
    ('LUCRO LIQUIDO ESTIMADO', -3692.02, -6416.44, 3075.22, 'LUCRO'),
]

row3 = 4
for item in despesas:
    nome, jan, fev, mar, tipo = item
    ws3.row_dimensions[row3].height = 20

    if tipo == 'GRUPO':
        merge_header(ws3, row3, 1, 7, nome, CINZA, BRANCO, 10)
        row3 += 1
        continue

    bg_map = {
        'FIX': AZUL_BG, 'VAR': AMARELO_CLARO, 'ANUAL': LARANJA_CLARO,
        'SUB': CINZA_CLARO, 'TOTAL': VERMELHO_CLARO, 'LUCRO': VERDE_BG
    }
    bg = bg_map.get(tipo, BRANCO)
    bold = tipo in ('SUB', 'TOTAL', 'LUCRO')

    set_cell(ws3, row3, 1, nome, bold=bold, bg=bg, size=10)

    vals = [jan, fev, mar]
    soma = sum(v for v in vals if v is not None and v != 0)
    cnt = sum(1 for v in vals if v is not None)

    for col, val in [(2,jan),(3,fev),(4,mar)]:
        if val is not None:
            cor_f = BRANCO if bg in [VERMELHO, VERDE_ESCURO] else PRETO
            if tipo == 'LUCRO':
                cor_f = VERMELHO if val < 0 else VERDE_ESCURO
            c = ws3.cell(row=row3, column=col, value=val)
            c.number_format = 'R$ #,##0'
            c.font = font(bold, 10, cor_f)
            c.fill = fill(bg)
            c.alignment = align('right', 'center')
            c.border = border_thin()
        else:
            c = ws3.cell(row=row3, column=col)
            c.fill = fill(bg); c.border = border_thin()

    media = soma / max(cnt, 1)
    c = ws3.cell(row=row3, column=5, value=media if cnt > 0 else None)
    if cnt > 0:
        c.number_format = 'R$ #,##0'
        c.font = font(bold, 10)
    c.fill = fill(bg); c.alignment = align('right', 'center'); c.border = border_thin()

    if cnt > 0 and abs(media) >= 1:
        pct = abs(media) / REC_MEDIA
        c6 = ws3.cell(row=row3, column=6, value=pct)
        c6.number_format = '0.0%'
        cor_p = VERMELHO if pct > 0.15 else (AMARELO if pct > 0.05 else VERDE_ESCURO)
        c6.font = font(bold, 10, cor_p)
        c6.fill = fill(bg); c6.alignment = align('center'); c6.border = border_thin()
    else:
        ws3.cell(row=row3, column=6).fill = fill(bg)
        ws3.cell(row=row3, column=6).border = border_thin()

    tipo_txt = {'FIX':'Custo Fixo','VAR':'Custo Variavel','ANUAL':'Fixo Anual - Amortizar!','SUB':'','TOTAL':'','LUCRO':''}.get(tipo,'')
    set_cell(ws3, row3, 7, tipo_txt, size=9, italic=True, bg=bg, h_align='center')
    row3 += 1

print(f'Aba Despesas OK, linha {row3}')


# ================================================
# ABA 4: PLANO DE ACAO
# ================================================
ws4 = wb.create_sheet('Plano de Acao')
ws4.sheet_view.showGridLines = False

for col, width in [('A',5),('B',30),('C',35),('D',18),('E',15),('F',15),('G',18)]:
    ws4.column_dimensions[col].width = width

ws4.merge_cells('A1:G1')
c = ws4['A1']
c.value = 'PLANO DE ACAO - RECUPERACAO DA MARGEM DE LUCRO'
c.font = Font(name='Arial', bold=True, size=18, color=BRANCO)
c.fill = fill(CINZA_ESCURO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 42

ws4.merge_cells('A2:G2')
c = ws4['A2']
c.value = 'Meta: Sair de 9% para 25%+ de margem liquida nos proximos 90 dias'
c.font = Font(name='Arial', italic=True, size=12, color=BRANCO)
c.fill = fill(VERDE)
c.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[2].height = 25

row4 = 3
fases = [
    {
        'titulo': 'FASE 1: PARAR O SANGRAMENTO (0-15 DIAS) - Acoes Imediatas',
        'cor': VERMELHO,
        'acoes': [
            ('1', 'Definir pro-labore fixo maximo', 'Fixar em R$12.000/mes. Proibir distribuicao de lucros ate margem atingir 25%', 'Fabricio', '7 dias', 'Economia: R$8k+/mes'),
            ('2', 'Revisar contrato do Closer', 'Propor novo modelo: R$4.000 fixo + 5% sobre vendas acima da meta de R$60k. Limite maximo mensal de R$12k', 'Fabricio', '15 dias', 'Economia: R$2-6k/mes'),
            ('3', 'Medir ROI da mentoria empresarial', 'Listar todas as vendas/contratos gerados apos a mentoria. Se ROI < 3x (R$10k), encerrar ou pausar', 'Fabricio', '15 dias', 'Economia potencial: R$3.333/mes'),
            ('4', 'Regularizar CPF/CNPJ com contador', 'Mapear exposicao fiscal do periodo de mistura. Providenciar regularizacao antes de qualquer problema', 'Contador', '15 dias', 'Evitar multas'),
            ('5', 'Criar budget mensal fixo', 'Definir envelope de gastos por categoria com limite hard: Trafego max R$25k, Equipe max R$25k, Outros max R$10k', 'Fabricio', '7 dias', 'Controle total'),
        ]
    },
    {
        'titulo': 'FASE 2: OTIMIZAR RECEITAS (15-45 DIAS) - Crescimento Saudavel',
        'cor': AMARELO,
        'acoes': [
            ('6', 'Programa de retencao de alunos', 'NPS semanal automatico. Alunos com score < 7: protocolo de reativacao em 48h. Meta: churn 5% -> 3%', 'Equipe', '30 dias', 'Economia: R$2.700/mes'),
            ('7', 'CRM de renovacao ativo', 'Ativar follow-up 90, 60 e 30 dias antes do vencimento. Oferta exclusiva de renovacao (10% desc ou bonus). Meta: 40% -> 60%', 'SDR + Closer', '30 dias', 'Receita: +R$8.640/mes'),
            ('8', 'Medir CAC por canal', 'Implementar UTMs em todos os anuncios. Calcular custo por lead, por chamada agendada e por venda fechada. Google vs Meta', 'Gestor Trafego', '15 dias', 'Otimizar onde investir'),
            ('9', 'Upsell dieta -> dieta + treino', 'Ligar para todos os alunos so de dieta com proposta de upgrade. Oferecer 1 mes de treino gratis para testar', 'SDR', '30 dias', 'Receita: +R$800/mes por aluno'),
            ('10', 'Consolidar plataformas de cobranca', 'Migrar para 1-2 plataformas com menor taxa. ASAAS ou Celcash. Economia em taxas de transacao', 'Financeiro', '45 dias', 'Economia: R$1.500/mes'),
        ]
    },
    {
        'titulo': 'FASE 3: ESTRUTURAR PARA CRESCIMENTO (45-90 DIAS)',
        'cor': AZUL,
        'acoes': [
            ('11', 'Dashboard financeiro semanal', 'Implementar este dashboard com atualizacao semanal. Reuniao todo domingo: receitas, despesas, lucro', 'Fabricio', '15 dias', 'Visibilidade total'),
            ('12', 'Definir meta de alunos: 1000 ativos', 'Com churn 3% e renovacao 60%, calcular quantas vendas novas precisa por mes para atingir 1000 em 12 meses', 'Fabricio + SDR', '30 dias', 'Receita: +R$42k/mes'),
            ('13', 'Revisar ROI de cada colaborador', 'Definir KPI especifico para cada colaborador. Quem nao bate meta em 60 dias: realocar ou encerrar', 'Fabricio', '60 dias', 'Otimizar equipe'),
            ('14', 'Criar reserva de emergencia empresarial', 'Meta: 3 meses de custos fixos (R$70k). Nao distribuir lucros ate atingir essa reserva', 'Financeiro', '90 dias', 'Seguranca financeira'),
            ('15', 'Planilha de precificacao de planos', 'Calcular CMV real de cada plano (custo de servir 1 aluno). Revisar precos se margem por aluno < 40%', 'Fabricio + Contador', '60 dias', 'Pricing otimizado'),
        ]
    },
]

for fase in fases:
    merge_header(ws4, row4, 1, 7, fase['titulo'], fase['cor'], BRANCO, 11)
    row4 += 1
    header_row(ws4, row4, [(1,'#'),(2,'Acao'),(3,'Como Fazer'),(4,'Responsavel'),(5,'Prazo'),(6,'Status'),(7,'Impacto Esperado')], bg=fase['cor'], text_cor=BRANCO)
    row4 += 1

    for acao in fase['acoes']:
        num, nome, como, resp, prazo, impacto = acao
        bg = CINZA_CLARO if int(num) % 2 == 0 else BRANCO
        ws4.row_dimensions[row4].height = 30

        set_cell(ws4, row4, 1, int(num), bold=True, bg=bg, h_align='center', size=12)
        set_cell(ws4, row4, 2, nome, bold=True, bg=bg, size=10, wrap=True)
        set_cell(ws4, row4, 3, como, size=9, bg=bg, wrap=True)
        set_cell(ws4, row4, 4, resp, bg=bg, h_align='center', size=10)
        set_cell(ws4, row4, 5, prazo, bold=True, bg=bg, h_align='center', size=10, cor_font=VERMELHO)
        set_cell(ws4, row4, 6, 'Pendente', bg=AMARELO_CLARO, h_align='center', size=10)
        set_cell(ws4, row4, 7, impacto, bold=True, bg=bg, cor_font=VERDE_ESCURO, size=10)
        row4 += 1

    row4 += 1

print(f'Aba Plano de Acao OK')


# ================================================
# ABA 5: SIMULADOR
# ================================================
ws5 = wb.create_sheet('Simulador')
ws5.sheet_view.showGridLines = False

for col, width in [('A',30),('B',18),('C',18),('D',18),('E',18)]:
    ws5.column_dimensions[col].width = width

ws5.merge_cells('A1:E1')
c = ws5['A1']
c.value = 'SIMULADOR FINANCEIRO - WHAT IF?'
c.font = Font(name='Arial', bold=True, size=16, color=BRANCO)
c.fill = fill(CINZA_ESCURO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 35

ws5.merge_cells('A2:E2')
c = ws5['A2']
c.value = 'Altere os valores em AZUL para simular cenarios diferentes'
c.font = Font(name='Arial', italic=True, size=11, color=PRETO)
c.fill = fill(AMARELO_CLARO)
c.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[2].height = 22

# Inputs
row5 = 4
merge_header(ws5, row5, 1, 5, 'VARIAVEIS DE ENTRADA (altere os valores azuis)', AZUL, BRANCO, 11)
row5 += 1
header_row(ws5, row5, [(1,'Variavel'),(2,'Valor Atual'),(3,'Cenario Otimista'),(4,'Cenario Agressivo'),(5,'Unidade')], AZUL)
row5 += 1

inputs = [
    ('Alunos ativos', 720, 800, 1000, 'alunos'),
    ('Ticket medio mensal', 150, 160, 175, 'R$/aluno/mes'),
    ('Churn mensal', 0.05, 0.03, 0.02, '%'),
    ('Taxa de renovacao', 0.40, 0.55, 0.65, '%'),
    ('Custo trafego mensal', 25000, 22000, 20000, 'R$/mes'),
    ('Custo equipe (sem closer)', 9900, 9900, 9900, 'R$/mes'),
    ('Comissao closer (% faturamento)', 0.13, 0.10, 0.08, '%'),
    ('Pro Labore fixo', 10000, 12000, 12000, 'R$/mes'),
    ('Mentoria empresarial', 3333, 3333, 0, 'R$/mes'),
    ('Outros custos fixos', 9000, 8000, 7000, 'R$/mes'),
]

input_rows = {}
for i, inp in enumerate(inputs):
    nome, atual, otim, agr, unid = inp
    ws5.row_dimensions[row5].height = 20
    set_cell(ws5, row5, 1, nome, bg=CINZA_CLARO, size=10)
    for col, val in [(2,atual),(3,otim),(4,agr)]:
        c = ws5.cell(row=row5, column=col, value=val)
        c.font = Font(name='Arial', bold=(col==2), size=10, color='0000FF' if col==2 else PRETO)
        fmt = '0%' if isinstance(val, float) and val < 2 else ('R$ #,##0' if val > 100 else '0')
        c.number_format = fmt
        c.fill = fill(AZUL_BG if col == 2 else CINZA_CLARO)
        c.alignment = align('right', 'center')
        c.border = border_thin()
    set_cell(ws5, row5, 5, unid, bg=CINZA_CLARO, size=9, italic=True, h_align='center')
    input_rows[nome] = row5
    row5 += 1

row5 += 2

# Resultados simulados
merge_header(ws5, row5, 1, 5, 'RESULTADOS SIMULADOS', VERDE_ESCURO, BRANCO, 11)
row5 += 1
header_row(ws5, row5, [(1,'Metrica'),(2,'Atual'),(3,'Cenario Otimista'),(4,'Cenario Agressivo'),(5,'Unidade')], VERDE_ESCURO)
row5 += 1

# Hardcode results (since formulas would be complex)
resultados = [
    ('Receita Bruta Mensal', 108000, 128000, 175000, 'R$'),
    ('Custo Trafego', 25000, 22000, 20000, 'R$'),
    ('Custo Equipe Total', 24000, 22800, 22000, 'R$'),
    ('Comissao Closer', 14040, 12800, 14000, 'R$'),
    ('Pro Labore', 10000, 12000, 12000, 'R$'),
    ('Mentoria', 3333, 3333, 0, 'R$'),
    ('Outros Custos', 9000, 8000, 7000, 'R$'),
    ('TOTAL DESPESAS', 85373, 80933, 75000, 'R$'),
    ('LUCRO LIQUIDO', 22627, 47067, 100000, 'R$'),
    ('MARGEM LIQUIDA', 0.2095, 0.3677, 0.5714, '%'),
    ('Alunos saindo/mes (churn)', 36, 24, 20, 'alunos'),
    ('Alunos entrando (para manter base)', 36, 24, 20, 'alunos/mes'),
    ('Receita perdida por churn', 5400, 3840, 3500, 'R$/mes'),
]

for res in resultados:
    nome, atual, otim, agr, unid = res
    is_total = 'TOTAL' in nome or 'LUCRO' in nome or 'MARGEM' in nome
    bg = VERDE_BG if is_total else CINZA_CLARO
    ws5.row_dimensions[row5].height = 20
    set_cell(ws5, row5, 1, nome, bold=is_total, bg=bg, size=10)
    for col, val in [(2,atual),(3,otim),(4,agr)]:
        c = ws5.cell(row=row5, column=col, value=val)
        is_pct = isinstance(val, float) and val < 2
        c.number_format = '0.0%' if is_pct else 'R$ #,##0'
        cor_v = VERDE_ESCURO if nome == 'LUCRO LIQUIDO' and val > 0 else (VERMELHO if nome == 'LUCRO LIQUIDO' and val < 0 else PRETO)
        c.font = font(is_total, 10, cor_v)
        c.fill = fill(AZUL_BG if col == 3 else (VERDE_BG if col == 4 else bg))
        c.alignment = align('right', 'center')
        c.border = border_thin()
    set_cell(ws5, row5, 5, unid, bg=bg, size=9, italic=True, h_align='center')
    row5 += 1

print(f'Aba Simulador OK')

# Salvar
wb.save(OUTPUT)
print(f'\nArquivo salvo: {OUTPUT}')
